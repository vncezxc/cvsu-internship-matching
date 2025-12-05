from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from django.core.mail import send_mail
import json
import csv
import openpyxl
from django.core.files.base import ContentFile
from io import BytesIO
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from accounts.models import User, StudentProfile, Course, RequiredDocument, AdviserProfile, CoordinatorProfile, DTR, StudentDocument
from internship.models import Company, Internship, Application
from .models import DashboardStatistics, Report
from .forms import RequiredDocumentForm, StudentDocumentUploadForm, DTRSubmissionForm
from .student_document_forms import StudentDocumentUploadForm as StudentDocumentUploadFormAlt
from django.views.decorators.http import require_POST, require_http_methods


# Student: Download templates and upload filled documents
@login_required
def student_documents_view(request):
    if not request.user.is_student:
        messages.error(request, 'Only students can access this page.')
        return redirect('dashboard:home')
    student = request.user.student_profile
    required_docs = RequiredDocument.objects.all()
    student_docs = StudentDocument.objects.filter(student=student)
    doc_map = {doc.document_type_id: doc for doc in student_docs}
    if request.method == 'POST':
        doc_id = request.POST.get('doc_id')
        file = request.FILES.get('file')
        required_doc = RequiredDocument.objects.filter(id=doc_id).first()
        if required_doc and file:
            stu_doc, created = StudentDocument.objects.get_or_create(student=student, document_type=required_doc)
            stu_doc.file = file
            stu_doc.approved = False  # Reset approval on new upload
            stu_doc.save()
            messages.success(request, f"Uploaded {required_doc.name}.")
        return redirect('dashboard:student_documents')
    return render(request, 'dashboard/student_documents.html', {
        'required_docs': required_docs,
        'doc_map': doc_map,
    })
# Adviser: Review student documents for their advisees
from django.views.decorators.http import require_POST

@login_required
def adviser_documents_review(request):
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can review student documents.')
        return redirect('dashboard:home')
    adviser = request.user.adviser_profile
    # Get advisees for this adviser: students in courses and sections handled by this adviser
    advisees = StudentProfile.objects.filter(
        course__in=adviser.courses.all(),
        section__in=adviser.get_sections_list()
    )
    required_docs = RequiredDocument.objects.all()
    student_docs = StudentDocument.objects.filter(student__in=advisees)
    # Map: {student_id: {doc_id: StudentDocument}}
    doc_map = {}
    for doc in student_docs:
        doc_map.setdefault(doc.student_id, {})[doc.document_type_id] = doc
    if request.method == 'POST':
        # Approve a document
        doc_id = request.POST.get('doc_id')
        stu_doc = StudentDocument.objects.filter(id=doc_id, student__in=advisees).first()
        if stu_doc:
            stu_doc.approved = True
            stu_doc.save()
            messages.success(request, f"Document '{stu_doc.document_type.name}' for {stu_doc.student.get_full_name()} approved.")
        return redirect('dashboard:adviser_documents_review')
    return render(request, 'dashboard/adviser_documents_review.html', {
        'advisees': advisees,
        'required_docs': required_docs,
        'doc_map': doc_map,
    })

# --- Coordinator: Upload/Update student document ---
@login_required
@require_http_methods(["GET", "POST"])
def coordinator_upload_student_document(request, student_id, doc_id):
    if not request.user.is_coordinator:
        messages.error(request, 'Only coordinators can upload/update student documents.')
        return redirect('dashboard:home')
    student = get_object_or_404(StudentProfile, id=student_id)
    required_doc = get_object_or_404(RequiredDocument, id=doc_id)
    existing = StudentDocument.objects.filter(student=student, document_type=required_doc).first()
    if request.method == 'POST':
        form = StudentDocumentUploadForm(request.POST, request.FILES, instance=existing)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.student = student
            doc.document_type = required_doc
            doc.save()
            messages.success(request, f'{required_doc.name} uploaded/updated for {student.get_full_name()}.')
            return redirect('dashboard:coordinator_documents_overview')
    else:
        form = StudentDocumentUploadForm(instance=existing)
    return render(request, 'dashboard/coordinator_upload_student_document.html', {'form': form, 'student': student, 'required_doc': required_doc})

# --- DTR Submission (Student) ---
@login_required
@require_http_methods(["GET", "POST"])
def submit_dtr(request):
    if not request.user.is_student:
        messages.error(request, 'Only students can submit DTRs.')
        return redirect('dashboard:home')
    profile = request.user.student_profile
    if request.method == 'POST':
        form = DTRSubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            dtr = form.save(commit=False)
            dtr.student = profile
            dtr.adviser = profile.course.advisers.first() if profile.course else None
            dtr.save()
            messages.success(request, 'DTR submitted for review.')
            return redirect('dashboard:student_dashboard')
    else:
        form = DTRSubmissionForm()
    return render(request, 'dashboard/submit_dtr.html', {'form': form})

# --- DTR List & Review (Adviser) ---
@login_required
def adviser_dtr_list(request):
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can view DTRs.')
        return redirect('dashboard:home')
    adviser = request.user.adviser_profile
    dtrs = DTR.objects.filter(adviser=adviser).order_by('-week_start')
    return render(request, 'dashboard/adviser_dtr_list.html', {'dtrs': dtrs})

@login_required
@require_http_methods(["GET", "POST"])
def review_dtr(request, dtr_id):
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can review DTRs.')
        return redirect('dashboard:home')
    dtr = get_object_or_404(DTR, id=dtr_id)
    if dtr.adviser != request.user.adviser_profile:
        messages.error(request, 'You do not have access to this DTR.')
        return redirect('dashboard:adviser_dtr_list')
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')
        if action == 'approve':
            dtr.approved = True
            dtr.remarks = remarks
            dtr.save()
            # Add hours to student only if not already counted
            student = dtr.student
            if not hasattr(dtr, '_hours_counted') or not dtr._hours_counted:
                student.ojt_hours_completed += dtr.hours_rendered
                student.save()
                dtr._hours_counted = True
                dtr.save()
            messages.success(request, 'DTR approved and hours credited.')
        elif action == 'reject':
            dtr.approved = False
            dtr.remarks = remarks or 'Rejected by adviser.'
            dtr.save()
            messages.info(request, 'DTR rejected.')
        else:
            messages.error(request, 'Invalid action.')
        return redirect('dashboard:adviser_dtr_list')
    return render(request, 'dashboard/review_dtr.html', {'dtr': dtr})
from django.db.models import Count
from django.core.mail import send_mail
import json
import csv
import openpyxl
from django.core.files.base import ContentFile
from io import BytesIO
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from accounts.models import User, StudentProfile, Course, RequiredDocument, AdviserProfile, CoordinatorProfile
from internship.models import Company, Internship, Application
from .models import DashboardStatistics, Report
from .forms import RequiredDocumentForm
from .student_document_forms import StudentDocumentUploadForm
from django.views.decorators.http import require_POST
from accounts.models import StudentDocument



@login_required
@require_POST
def review_document(request, document_id):
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can review documents.')
        return redirect('dashboard:home')
    document = get_object_or_404(StudentDocument, id=document_id)
    adviser = request.user.adviser_profile
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '')
    if action == 'accept':
        document.status = 'ACCEPTED'
        document.adviser = adviser
        document.adviser_remarks = remarks
        document.save()
        messages.success(request, 'Document accepted and status updated.')
    elif action == 'reject':
        document.status = 'REJECTED'
        document.adviser = adviser
        document.adviser_remarks = remarks or 'Rejected by adviser.'
        document.save()
        messages.info(request, 'Document rejected.')
    else:
        messages.error(request, 'Invalid action.')
    return redirect('dashboard:student_detail', student_id=document.student.id)

# Dashboard redirect
@login_required
def dashboard_redirect(request):
    """Redirect to appropriate dashboard based on user type."""
    user = request.user
    
    if user.is_student:
        return redirect('dashboard:student_dashboard')
    elif user.is_adviser:
        return redirect('dashboard:adviser_dashboard')
    elif user.is_coordinator:
        return redirect('dashboard:coordinator_dashboard')
    else:
        messages.warning(request, 'You do not have access to any dashboard.')
        return redirect('home')

# Student dashboard
@login_required
def student_dashboard(request):
    """Dashboard for student users."""
    if not request.user.is_student:
        messages.error(request, 'You do not have access to the student dashboard.')
        return redirect('dashboard:home')
    
    try:
        profile = request.user.student_profile
        
        # Get application statistics
        applications = Application.objects.filter(student=profile)
        pending_count = applications.filter(status=Application.Status.PENDING).count()
        accepted_count = applications.filter(status=Application.Status.ACCEPTED).count()
        rejected_count = applications.filter(status=Application.Status.REJECTED).count()
        
        # Get recent applications
        recent_applications = applications.order_by('-applied_at')[:5]
        
        # Check profile completeness
        # Check profile completeness for matching (course, skills, pinned location)
        if not profile.profile_is_complete_for_matching:
            messages.warning(request, 'Please complete your profile to see internship matches.')
        
        # Document completion logic (including CV/Resume)
        from accounts.models import RequiredDocument, StudentDocument
        required_documents = list(RequiredDocument.objects.all())
        uploaded_documents = {doc.document_type_id: doc for doc in profile.documents.all()}
        document_status = []
        for req_doc in required_documents:
            doc = uploaded_documents.get(req_doc.id)
            document_status.append({
                'required': req_doc,
                'uploaded': doc is not None,
                'file': doc.file.url if doc else None,
                'doc_id': doc.id if doc else None,
            })

        # Add CV/Resume as a pseudo-required document
        cv_entry = {
            'required': type('obj', (object,), {
                'name': 'CV/Resume',
                'description': 'Student CV or Resume',
                'id': 'cv',
            })(),
            'uploaded': bool(profile.cv),
            'file': profile.cv.url if profile.cv else None,
            'doc_id': None,
        }
        document_status.append(cv_entry)

        # Calculate document completion percent (including CV)
        total_docs = len(document_status)
        completed_docs = len([d for d in document_status if d['uploaded']])
        if total_docs > 0:
            document_completion_percent = int((completed_docs / total_docs) * 100)
        else:
            document_completion_percent = 0
        context = {
            'profile': profile,
            'application_count': applications.count(),
            'pending_count': pending_count,
            'accepted_count': accepted_count,
            'rejected_count': rejected_count,
            'recent_applications': recent_applications,
            'progress_percentage': profile.get_progress_percentage(),
            'profile_completion_percentage': profile.get_profile_completion_percentage(),
            'document_status': document_status,
            'document_completion_percent': document_completion_percent,
        }
        return render(request, 'dashboard/student_dashboard.html', context)
    except StudentProfile.DoesNotExist:
        messages.warning(request, 'Please complete your profile first.')
        return redirect('accounts:edit_profile')

# Adviser dashboard
@login_required
def adviser_dashboard(request):
    """Dashboard for OJT adviser users."""
    if not request.user.is_adviser:
        messages.error(request, 'You do not have access to the adviser dashboard.')
        return redirect('dashboard:home')
    # Auto-create adviser profile if missing
    try:
        profile = request.user.adviser_profile
        print(f"DEBUG: AdviserProfile found for user {request.user.id} ({request.user.email})")
    except AdviserProfile.DoesNotExist:
        print(f"DEBUG: AdviserProfile NOT found for user {request.user.id} ({request.user.email})")
        course = Course.objects.first()
        profile = AdviserProfile.objects.create(user=request.user, department='IT Department')
        if course:
            profile.courses.set([course])
        profile.sections = 'A,B'
        profile.save()
    
    # Get students handled by this adviser
    students = StudentProfile.objects.filter(
        course__in=profile.courses.all(),
        section__in=profile.get_sections_list()
    )
    
    # Get student statistics
    total_students = students.count()
    looking_count = students.filter(ojt_status=StudentProfile.OJTStatus.LOOKING).count()
    waiting_count = students.filter(ojt_status=StudentProfile.OJTStatus.WAITING).count()
    ongoing_count = students.filter(ojt_status=StudentProfile.OJTStatus.ONGOING).count()
    completed_count = students.filter(ojt_status=StudentProfile.OJTStatus.COMPLETED).count()
    
    # Get recent students
    recent_students = students.order_by('-user__date_joined')[:5]
    
    # Document completion status
    from accounts.models import RequiredDocument, StudentDocument
    required_documents = list(RequiredDocument.objects.all())
    students_with_docs = []
    for student in students:
        uploaded_docs = {doc.document_type_id: doc for doc in student.documents.all()}
        doc_status = []
        for req_doc in required_documents:
            doc = uploaded_docs.get(req_doc.id)
            doc_status.append({
                'required': req_doc,
                'uploaded': doc is not None,
                'file': doc.file.url if doc else None,
            })
        # Progress calculation
        total = len(required_documents)
        completed = sum(1 for d in doc_status if d['uploaded'])
        progress = int((completed / total) * 100) if total > 0 else 0
        students_with_docs.append({
            'student': student,
            'doc_status': doc_status,
            'progress': progress,
        })
    context = {
        'profile': profile,
        'total_students': total_students,
        'looking_count': looking_count,
        'waiting_count': waiting_count,
        'ongoing_count': ongoing_count,
        'completed_count': completed_count,
        'recent_students': recent_students,
        'students_with_docs': students_with_docs,
        'required_documents': required_documents,
    }
    return render(request, 'dashboard/adviser_dashboard.html', context)

# Coordinator dashboard
@login_required
def coordinator_dashboard(request):
    """Dashboard for OJT coordinator users."""
    if not request.user.is_coordinator:
        messages.error(request, 'You do not have access to the coordinator dashboard.')
        return redirect('dashboard:home')
    # Auto-create coordinator profile if missing
    try:
        profile = request.user.coordinator_profile
        print(f"DEBUG: CoordinatorProfile found for user {request.user.id} ({request.user.email})")
    except CoordinatorProfile.DoesNotExist:
        print(f"DEBUG: CoordinatorProfile NOT found for user {request.user.id} ({request.user.email})")
        profile = CoordinatorProfile.objects.create(user=request.user, department='OJT Office')
    
    profile = request.user.coordinator_profile
    
    # Get or update statistics
    stats = DashboardStatistics.update_statistics()
    
    # Get recent companies and internships
    recent_companies = Company.objects.order_by('-created_at')[:5]
    recent_internships = Internship.objects.order_by('-created_at')[:5]
    
    # Get recent applications
    recent_applications = Application.objects.order_by('-applied_at')[:5]
    
    context = {
        'profile': profile,
        'stats': stats,
        'recent_companies': recent_companies,
        'recent_internships': recent_internships,
        'recent_applications': recent_applications,
    }
    return render(request, 'dashboard/coordinator_dashboard.html', context)

# Student management
@login_required
def student_list(request):
    """View list of students for advisers."""
    if not request.user.is_adviser:
        messages.error(request, 'Only OJT Advisers can access student management.')
        return redirect('dashboard:home')
    # Auto-create adviser profile if missing
    from accounts.models import AdviserProfile, Course
    try:
        profile = request.user.adviser_profile
    except AdviserProfile.DoesNotExist:
        course = Course.objects.first()
        profile = AdviserProfile.objects.create(user=request.user, department='IT Department')
        if course:
            profile.courses.set([course])
        profile.sections = 'A,B'
        profile.save()
    # Get students handled by this adviser
    students = StudentProfile.objects.filter(
        course__in=profile.courses.all(),
        section__in=profile.get_sections_list()
    ).select_related('user', 'course').prefetch_related('documents').order_by('user__last_name')
    
    # Filter by course if specified
    course_id = request.GET.get('course')
    if course_id:
        students = students.filter(course_id=course_id)
    
    # Filter by section if specified
    section = request.GET.get('section')
    if section:
        students = students.filter(section=section)
    
    # Filter by status if specified
    status = request.GET.get('status')
    if status in [s[0] for s in StudentProfile.OJTStatus.choices]:
        students = students.filter(ojt_status=status)
    
    # Compute requirements completion per student (RequiredDocument + CV)
    from accounts.models import RequiredDocument
    required_documents = list(RequiredDocument.objects.all())

    # Convert queryset to list so we can attach computed attrs
    students_list = list(students)
    for s in students_list:
        # collect uploaded document type ids for this student
        uploaded_doc_type_ids = {d.document_type_id for d in s.documents.all()}
        missing = [rd.name for rd in required_documents if rd.id not in uploaded_doc_type_ids]
        # include CV/Resume as required
        if not s.cv:
            missing.append('CV/Resume')
        s.requirements_complete = (len(missing) == 0)
        s.missing_requirements = missing

    context = {
        'students': students_list,
        'courses': profile.courses.all(),
        'sections': profile.get_sections_list(),
        'course_filter': course_id,
        'section_filter': section,
        'status_filter': status,
    }
    return render(request, 'dashboard/student_list.html', context)

@login_required
def student_detail(request, student_id):
    """View student details for advisers."""
    if not request.user.is_adviser:
        messages.error(request, 'Only OJT Advisers can access student details.')
        return redirect('dashboard:home')
    
    student = get_object_or_404(StudentProfile, id=student_id)
    
    # Check if adviser has access to this student
    profile = request.user.adviser_profile
    if not (student.course in profile.courses.all() and student.section in profile.get_sections_list()):
        messages.error(request, 'You do not have access to this student.')
        return redirect('dashboard:students')
    
    # Get applications for this student
    applications = Application.objects.filter(student=student).order_by('-applied_at')
    
    # Profile completion percentage
    profile_completion_percentage = student.get_profile_completion_percentage()

    # Document completion logic (including CV/Resume)
    from accounts.models import RequiredDocument, StudentDocument
    required_documents = list(RequiredDocument.objects.all())
    uploaded_documents = {doc.document_type_id: doc for doc in student.documents.all()}
    document_status = []
    for req_doc in required_documents:
        doc = uploaded_documents.get(req_doc.id)
        document_status.append({
            'required': req_doc,
            'uploaded': doc is not None,
            'file': doc.file.url if doc else None,
            'doc_id': doc.id if doc else None,
        })
    # Add CV/Resume as a pseudo-required document
    cv_entry = {
        'required': type('obj', (object,), {
            'name': 'CV/Resume',
            'description': 'Student CV or Resume',
            'id': 'cv',
        })(),
        'uploaded': bool(student.cv),
        'file': student.cv.url if student.cv else None,
        'doc_id': None,
    }
    document_status.append(cv_entry)
    # Calculate document completion percent (including CV)
    total_docs = len(document_status)
    completed_docs = len([d for d in document_status if d['uploaded']])
    if total_docs > 0:
        document_completion_percent = int((completed_docs / total_docs) * 100)
    else:
        document_completion_percent = 0
    # List of missing documents
    missing_documents = [d['required'].name for d in document_status if not d['uploaded']]

    context = {
        'student': student,
        'applications': applications,
        'progress_percentage': student.get_progress_percentage(),
        'profile_completion_percentage': profile_completion_percentage,
        'document_status': document_status,
        'document_completion_percent': document_completion_percent,
        'missing_documents': missing_documents,
    }
    return render(request, 'dashboard/student_detail.html', context)

@login_required
def update_student_status(request, student_id):
    """Update student OJT status."""
    if not request.user.is_adviser:
        messages.error(request, 'Only OJT Advisers can update student status.')
        return redirect('dashboard:home')
    
    student = get_object_or_404(StudentProfile, id=student_id)
    
    # Check if adviser has access to this student
    profile = request.user.adviser_profile
    if not (student.course in profile.courses.all() and student.section in profile.get_sections_list()):
        messages.error(request, 'You do not have access to this student.')
        return redirect('dashboard:students')
    
    if request.method == 'POST':
        status = request.POST.get('status')
        if status in [s[0] for s in StudentProfile.OJTStatus.choices]:
            student.ojt_status = status
            student.save()
            messages.success(request, f'Status updated to {student.get_ojt_status_display()}.')
        else:
            messages.error(request, 'Invalid status.')
        return redirect('dashboard:student_detail', student_id=student.id)
    
    context = {
        'student': student,
        'statuses': StudentProfile.OJTStatus.choices,
    }
    return render(request, 'dashboard/update_student_status.html', context)

@login_required
def update_ojt_hours(request, student_id):
    """Update student OJT hours completed."""
    if request.user.is_coordinator:
        messages.error(request, 'OJT Coordinators cannot update OJT hours.')
        return redirect('dashboard:home')

    student = get_object_or_404(StudentProfile, id=student_id)


    # Only advisers can update OJT hours
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can update OJT hours.')
        return redirect('dashboard:home')

    # Adviser access check
    profile = request.user.adviser_profile
    if not (student.course in profile.courses.all() and student.section in profile.get_sections_list()):
        messages.error(request, 'You do not have access to this student.')
        return redirect('dashboard:students')

    # Only allow update if student status is ONGOING
    if student.ojt_status != StudentProfile.OJTStatus.ONGOING:
        messages.error(request, 'OJT hours can only be updated when student status is "Currently Undergoing OJT".')
        return redirect('dashboard:student_detail', student_id=student.id)

    if request.method == 'POST':
        try:
            hours = int(request.POST.get('hours', 0))
            if hours >= 0:
                student.ojt_hours_completed = hours
                student.save()
                messages.success(request, f'OJT hours updated to {hours}.')
                # Email notification to student
                send_mail(
                    subject='OJT Hours Updated',
                    message=f'Your OJT hours have been updated to {hours} by your adviser.',
                    from_email=None,
                    recipient_list=[student.user.email],
                    fail_silently=True,
                )
            else:
                messages.error(request, 'Hours cannot be negative.')
        except ValueError:
            messages.error(request, 'Invalid hours value.')
        return redirect('dashboard:student_detail', student_id=student.id)

    context = {
        'student': student,
    }
    return render(request, 'dashboard/update_ojt_hours.html', context)

# Reports
from .report_filter_forms import ReportFilterForm

@login_required
def reports_dashboard(request):
    """Dashboard for reports with filtering."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')

    form = ReportFilterForm(request.GET or None)
    reports_qs = Report.objects.all().order_by('-generated_at')

    if form.is_valid():
        report_type = form.cleaned_data.get('report_type')
        generated_by = form.cleaned_data.get('generated_by')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')

        if report_type:
            reports_qs = reports_qs.filter(report_type=report_type)
        if generated_by:
            reports_qs = reports_qs.filter(
                Q(generated_by__first_name__icontains=generated_by) |
                Q(generated_by__last_name__icontains=generated_by) |
                Q(generated_by__email__icontains=generated_by)
            )
        if date_from:
            reports_qs = reports_qs.filter(generated_at__date__gte=date_from)
        if date_to:
            reports_qs = reports_qs.filter(generated_at__date__lte=date_to)

    # Sort by course and section for coordinators, and for advisers by their own students
    if request.user.is_coordinator:
        recent_reports = list(reports_qs)
        def report_sort_key(report):
            # Try to extract course and section from the report file name or metadata if possible
            # This assumes the report filename contains course and section, otherwise fallback to generated_at
            fname = (report.file.name if report.file else '').lower()
            import re
            course = section = ''
            m = re.search(r'(bs[a-z]+)', fname)
            if m:
                course = m.group(1)
            m2 = re.search(r'section[_-]?([a-z0-9]+)', fname)
            if m2:
                section = m2.group(1)
            return (course, section, -report.generated_at.timestamp())
        recent_reports.sort(key=report_sort_key)
        recent_reports = recent_reports[:10]
    elif request.user.is_adviser:
        adviser = getattr(request.user, 'adviser_profile', None)
        recent_reports = list(reports_qs.filter(generated_by=request.user))
        def report_sort_key(report):
            fname = (report.file.name if report.file else '').lower()
            import re
            course = section = ''
            m = re.search(r'(bs[a-z]+)', fname)
            if m:
                course = m.group(1)
            m2 = re.search(r'section[_-]?([a-z0-9]+)', fname)
            if m2:
                section = m2.group(1)
            return (course, section, -report.generated_at.timestamp())
        recent_reports.sort(key=report_sort_key)
        recent_reports = recent_reports[:10]
    else:
        recent_reports = reports_qs[:10]

    context = {
        'recent_reports': recent_reports,
        'is_coordinator': request.user.is_coordinator,
        'filter_form': form,
    }
    return render(request, 'dashboard/reports_dashboard.html', context)

@login_required
def generate_student_list(request):
    """Generate student list report as Excel."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    # Filter students for advisers
    if request.user.is_adviser:
        adviser_profile = request.user.adviser_profile
        students = StudentProfile.objects.filter(
            course__in=adviser_profile.courses.all(),
            section__in=adviser_profile.get_sections_list()
        ).select_related('user', 'course')
    else:
        students = StudentProfile.objects.all().select_related('user', 'course')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append(['Student ID', 'Name', 'Email', 'Course', 'Section', 'OJT Status', 'OJT Hours'])
    for s in students:
        ws.append([
            s.student_id or '',
            s.get_full_name(),
            s.user.email,
            s.course.name if s.course else '',
            s.section,
            s.get_ojt_status_display(),
            s.ojt_hours_completed,
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    report = Report.objects.create(
        report_type=Report.ReportType.STUDENT_LIST,
        generated_by=request.user
    )
    report.file.save('student_list.xlsx', ContentFile(output.read()))
    messages.success(request, 'Student list report generated successfully.')
    return redirect('dashboard:reports')

@login_required
def generate_ojt_tracking(request):
    """Generate OJT tracking sheet report as Excel."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    # Filter students for advisers
    if request.user.is_adviser:
        adviser_profile = request.user.adviser_profile
        students = StudentProfile.objects.filter(
            course__in=adviser_profile.courses.all(),
            section__in=adviser_profile.get_sections_list()
        ).select_related('user', 'course')
    else:
        students = StudentProfile.objects.all().select_related('user', 'course')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OJT Tracking'
    ws.append(['Name', 'Course', 'Section', 'OJT Status', 'OJT Hours Completed', 'OJT Hours Required'])
    for s in students:
        ws.append([
            s.get_full_name(),
            s.course.name if s.course else '',
            s.section,
            s.get_ojt_status_display(),
            s.ojt_hours_completed,
            s.ojt_hours_required,
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    report = Report.objects.create(
        report_type=Report.ReportType.OJT_TRACKING,
        generated_by=request.user
    )
    report.file.save('ojt_tracking.xlsx', ContentFile(output.read()))
    messages.success(request, 'OJT tracking sheet generated successfully.')
    return redirect('dashboard:reports')

@login_required
def generate_student_list(request):
    """Generate student list report as Excel."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    # Filter students for advisers
    if request.user.is_adviser:
        adviser_profile = request.user.adviser_profile
        students = StudentProfile.objects.filter(
            course__in=adviser_profile.courses.all(),
            section__in=adviser_profile.get_sections_list()
        ).select_related('user', 'course')
    else:
        students = StudentProfile.objects.all().select_related('user', 'course')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append(['Student ID', 'Name', 'Email', 'Course', 'Section', 'Year Level', 'OJT Status', 'OJT Hours', 'Company', 'Company Email', 'HR Email', 'Internship Title'])
    for s in students:
        # Get latest accepted application (if any)
        application = s.applications.filter(status='ACCEPTED').order_by('-applied_at').first()
        company = application.internship.company if application and application.internship else None
        internship = application.internship if application else None
        ws.append([
            s.student_id or '',
            s.get_full_name(),
            s.user.email,
            s.course.name if s.course else '',
            s.section,
            s.get_year_level_display() if hasattr(s, 'get_year_level_display') else s.year_level,
            s.get_ojt_status_display(),
            s.ojt_hours_completed,
            company.name if company else '',
            company.company_email if company else '',
            company.hr_email if company else '',
            internship.title if internship else '',
        ])
    # Adjust column widths for visibility
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    report = Report.objects.create(
        report_type=Report.ReportType.STUDENT_LIST,
        generated_by=request.user
    )
    report.file.save('student_list.xlsx', ContentFile(output.read()))
    messages.success(request, 'Student list report generated successfully.')
    return redirect('dashboard:reports')

@login_required
def download_report(request, report_id):
    """Download a generated report."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    
    report = get_object_or_404(Report, id=report_id)
    
    # Check if file exists
    if not report.file:
        messages.error(request, 'Report file not found.')
        return redirect('dashboard:reports')
    
    # Return the file for download
    response = HttpResponse(report.file, content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{report.file.name}"'
    return response

# Statistics API
@login_required
def api_statistics(request):
    """API endpoint for dashboard statistics."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    # Get date range from request
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Get statistics for the date range
    stats = DashboardStatistics.objects.filter(date__range=[start_date, end_date]).order_by('date')
    
    # Format data for charts
    dates = [stat.date.strftime('%Y-%m-%d') for stat in stats]
    student_data = {
        'looking': [stat.students_looking for stat in stats],
        'waiting': [stat.students_waiting for stat in stats],
        'ongoing': [stat.students_ongoing for stat in stats],
        'completed': [stat.students_completed for stat in stats],
    }
    company_data = {
        'active': [stat.active_companies for stat in stats],
        'inactive': [stat.total_companies - stat.active_companies for stat in stats],
    }
    application_data = {
        'pending': [stat.pending_applications for stat in stats],
        'accepted': [stat.accepted_applications for stat in stats],
        'rejected': [stat.rejected_applications for stat in stats],
    }
    
    return JsonResponse({
        'dates': dates,
        'student_data': student_data,
        'company_data': company_data,
        'application_data': application_data,
    })

@login_required
def company_list(request):
    """List all companies for coordinators."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can manage companies.')
        return redirect('dashboard:home')
    companies = Company.objects.all().order_by('-created_at')
    return render(request, 'internship/company_list.html', {'companies': companies})

@login_required
def add_company(request):
    """Add a new company (coordinator only). Handles all fields and file upload."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can add companies.')
        return redirect('dashboard:home')
    if request.method == 'POST':
        name = request.POST.get('name')
        status = request.POST.get('status', 'ACTIVE')
        description = request.POST.get('description')
        company_email = request.POST.get('company_email')
        hr_email = request.POST.get('hr_email')
        phone_number = request.POST.get('phone_number')
        street = request.POST.get('street')
        barangay = request.POST.get('barangay')
        city = request.POST.get('city')
        province = request.POST.get('province')
        latitude = request.POST.get('latitude') or None
        longitude = request.POST.get('longitude') or None
        has_incentives = request.POST.get('has_incentives') == 'on'
        incentives_details = request.POST.get('incentives_details')
        logo = request.FILES.get('logo')
        # Create company
        if name and company_email:
            company = Company.objects.create(
                name=name,
                status=status,
                description=description,
                company_email=company_email,
                hr_email=hr_email,
                phone_number=phone_number,
                street=street,
                barangay=barangay,
                city=city,
                province=province,
                latitude=latitude,
                longitude=longitude,
                has_incentives=has_incentives,
                incentives_details=incentives_details,
                logo=logo,
            )
            messages.success(request, 'Company added successfully.')
            # Email notification to HR
            if hr_email:
                send_mail(
                    subject='Company Registered',
                    message=f'Your company {name} has been registered as an OJT partner.',
                    from_email=None,
                    recipient_list=[hr_email],
                    fail_silently=True,
                )
            return redirect('dashboard:company_list')
        else:
            messages.error(request, 'Name and company email are required.')
    return render(request, 'internship/add_company.html')

@login_required
def edit_company(request, company_id):
    """Edit an existing company (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can edit companies.')
        return redirect('dashboard:home')
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        company.name = request.POST.get('name', company.name)
        company.company_email = request.POST.get('company_email', company.company_email)
        company.hr_email = request.POST.get('hr_email', company.hr_email)
        # ... update other fields as needed ...
        company.save()
        messages.success(request, 'Company updated successfully.')
        return redirect('dashboard:company_list')
    return render(request, 'internship/edit_company.html', {'company': company})

@login_required
def toggle_company_status(request, company_id):
    """Activate/deactivate a company (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can change company status.')
        return redirect('dashboard:home')
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        company.is_active = not company.is_active
        company.save()
        messages.success(request, f'Company status changed to {"Active" if company.is_active else "Inactive"}.')
        # Email notification to company
        if company.company_email:
            send_mail(
                subject='Company Status Changed',
                message=f'Your company status is now {"Active" if company.is_active else "Inactive"}.',
                from_email=None,
                recipient_list=[company.company_email],
                fail_silently=True,
            )
        return redirect('dashboard:company_list')
    return render(request, 'internship/confirm_toggle_status.html', {'company': company})

@login_required
def edit_adviser_profile(request):
    """Edit adviser profile."""
    if not request.user.is_adviser:
        messages.error(request, 'Only advisers can edit their profile.')
        return redirect('dashboard:home')
    profile = request.user.adviser_profile
    if request.method == 'POST':
        # Example: update fields (expand as needed)
        profile.phone_number = request.POST.get('phone_number', profile.phone_number)
        # ... add more fields as needed ...
        profile.save()
        messages.success(request, 'Profile updated successfully.')
        # Email notification
        send_mail(
            subject='Profile Updated',
            message='Your adviser profile has been updated.',
            from_email=None,
            recipient_list=[request.user.email],
            fail_silently=True,
        )
        return redirect('dashboard:adviser_dashboard')
    return render(request, 'dashboard/edit_adviser_profile.html', {'profile': profile})

@login_required
def edit_coordinator_profile(request):
    """Edit coordinator profile."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only coordinators can edit their profile.')
        return redirect('dashboard:home')
    profile = request.user.coordinator_profile
    if request.method == 'POST':
        # Example: update fields (expand as needed)
        profile.phone_number = request.POST.get('phone_number', profile.phone_number)
        # ... add more fields as needed ...
        profile.save()
        messages.success(request, 'Profile updated successfully.')
        # Email notification
        send_mail(
            subject='Profile Updated',
            message='Your coordinator profile has been updated.',
            from_email=None,
            recipient_list=[request.user.email],
            fail_silently=True,
        )
        return redirect('dashboard:coordinator_dashboard')
    return render(request, 'dashboard/edit_coordinator_profile.html', {'profile': profile})

@login_required
def generate_student_list_pdf(request):
    """Generate student list report as PDF."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    # Filter students for advisers
    if request.user.is_adviser:
        adviser_profile = request.user.adviser_profile
        students = StudentProfile.objects.filter(
            course__in=adviser_profile.courses.all(),
            section__in=adviser_profile.get_sections_list()
        ).select_related('user', 'course')
    else:
        students = StudentProfile.objects.all().select_related('user', 'course')
    from reportlab.lib.pagesizes import A4
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    data = [['Student ID', 'Name', 'Email', 'Course', 'Section', 'Year Level', 'OJT Status', 'OJT Hours', 'Company', 'Company Email', 'HR Email', 'Internship Title']]
    for s in students:
        application = s.applications.filter(status='ACCEPTED').order_by('-applied_at').first()
        company = application.internship.company if application and application.internship else None
        internship = application.internship if application else None
        data.append([
            s.student_id or '',
            s.get_full_name(),
            s.user.email,
            s.course.code if s.course and hasattr(s.course, 'code') else (s.course.name if s.course else ''),
            s.section,
            s.get_year_level_display() if hasattr(s, 'get_year_level_display') else s.year_level,
            s.get_ojt_status_display(),
            s.ojt_hours_completed,
            company.name if company else '',
            company.company_email if company else '',
            company.hr_email if company else '',
            internship.title if internship else '',
        ])
    # Adjusted column widths for better fit
    # Adjusted columns and font for A4 fit
    col_widths = [45, 80, 90, 55, 35, 45, 83, 35, 65, 80, 80, 70]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (0,1), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('FONTSIZE', (0,1), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('TOPPADDING', (0,1), (-1,-1), 2),
        ('BOTTOMPADDING', (0,1), (-1,-1), 2),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ])
    # Alternating row background for readability
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0,i), (-1,i), colors.whitesmoke)
        else:
            style.add('BACKGROUND', (0,i), (-1,i), colors.beige)
    # Enable word wrap for all columns
    for col in range(len(data[0])):
        style.add('WORDWRAP', (col, 0), (col, -1), True)
    table.setStyle(style)
    # Center table on A3 page
    table_width, table_height = table.wrapOn(p, 0, 0)
    page_width, page_height = landscape(A4)
    x = (page_width - table_width) / 2
    y = page_height - table_height - 80
    table.drawOn(p, x, y)
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def generate_ojt_tracking_pdf(request):
    """Generate OJT tracking sheet report as PDF."""
    if not (request.user.is_adviser or request.user.is_coordinator):
        messages.error(request, 'You do not have access to reports.')
        return redirect('dashboard:home')
    students = StudentProfile.objects.all().select_related('user', 'course')
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    data = [['Name', 'Course', 'Section', 'OJT Status', 'OJT Hours Completed', 'OJT Hours Required']]
    for s in students:
        data.append([
            s.get_full_name(),
            s.course.name if s.course else '',
            s.section,
            s.get_ojt_status_display(),
            s.ojt_hours_completed,
            s.ojt_hours_required,
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    table.wrapOn(p, 800, 600)
    table.drawOn(p, 30, 500 - 20 * len(data))
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def generate_company_list_pdf(request):
    """Generate company list report as PDF."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can generate company reports.')
        return redirect('dashboard:home')
    companies = Company.objects.all()
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    data = [['Name', 'Email', 'HR Email', 'Phone', 'Status', 'City', 'Province']]
    for c in companies:
        data.append([
            c.name,
            c.company_email,
            c.hr_email,
            c.phone_number,
            c.status,
            c.city,
            c.province,
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    table.wrapOn(p, 800, 600)
    table.drawOn(p, 30, 500 - 20 * len(data))
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def generate_application_summary_pdf(request):
    """Generate application summary report as PDF."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can generate application reports.')
        return redirect('dashboard:home')
    applications = Application.objects.select_related('student', 'internship')
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    data = [['Student', 'Internship', 'Company', 'Status', 'Applied At']]
    for a in applications:
        data.append([
            a.student.get_full_name(),
            a.internship.title if a.internship else '',
            a.internship.company.name if a.internship and a.internship.company else '',
            a.get_status_display(),
            a.applied_at.strftime('%Y-%m-%d %H:%M'),
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    table.wrapOn(p, 800, 600)
    table.drawOn(p, 30, 500 - 20 * len(data))
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def required_documents_list(request):
    """List all required documents for coordinators."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can manage required documents.')
        return redirect('dashboard:home')
    documents = RequiredDocument.objects.all()
    return render(request, 'dashboard/required_documents_list.html', {'documents': documents})

@login_required
def add_required_document(request):
    """Add a new required document (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can add required documents.')
        return redirect('dashboard:home')
    if request.method == 'POST':
        form = RequiredDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save()
            # If DOCX template uploaded, generate initial CKEditor HTML
            print(f"DEBUG: template_file={doc.template_file}, name={getattr(doc.template_file, 'name', None)}")
            if doc.template_file and str(getattr(doc.template_file, 'name', '')).lower().endswith('.docx'):
                from dashboard.docx_utils_html import extract_docx_full_html
                import os
                from django.conf import settings
                docx_path = os.path.join(settings.MEDIA_ROOT, doc.template_file.name)
                html_output_path = os.path.join(settings.MEDIA_ROOT, f"required_doc_{doc.id}_full.html")
                html_content = extract_docx_full_html(docx_path)
                with open(html_output_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
            messages.success(request, 'Required document added.')
            return redirect('dashboard:required_documents_list')
        else:
            print(f"DEBUG: form.errors={form.errors}")
    else:
        form = RequiredDocumentForm()
    return render(request, 'dashboard/required_document_form.html', {'form': form, 'action': 'Add'})

@login_required
def edit_required_document(request, doc_id):
    """Edit an existing required document (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can edit required documents.')
        return redirect('dashboard:home')
    document = get_object_or_404(RequiredDocument, id=doc_id)
    if request.method == 'POST':
        form = RequiredDocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, 'Required document updated.')
            return redirect('dashboard:required_documents_list')
    else:
        form = RequiredDocumentForm(instance=document)
    return render(request, 'dashboard/required_document_form.html', {'form': form, 'action': 'Edit'})

@login_required
def delete_required_document(request, doc_id):
    """Delete a required document (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only OJT Coordinators can delete required documents.')
        return redirect('dashboard:home')
    document = get_object_or_404(RequiredDocument, id=doc_id)
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Required document deleted.')
        return redirect('dashboard:required_documents_list')
    return render(request, 'dashboard/required_document_confirm_delete.html', {'document': document})

@login_required
def upload_student_document(request, doc_id):
    """Upload a required document (student only)."""
    if not request.user.is_student:
        messages.error(request, 'Only students can upload documents.')
        return redirect('dashboard:home')
    from accounts.models import RequiredDocument, StudentDocument
    profile = request.user.student_profile
    required_doc = get_object_or_404(RequiredDocument, id=doc_id)
    if request.method == 'POST':
        form = StudentDocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Remove old document if exists
            StudentDocument.objects.filter(student=profile, document_type=required_doc).delete()
            doc = form.save(commit=False)
            doc.student = profile
            doc.document_type = required_doc
            doc.save()
            messages.success(request, f'{required_doc.name} uploaded successfully.')
            return redirect('dashboard:student_dashboard')
    else:
        form = StudentDocumentUploadForm(initial={'document_type': required_doc})
    return render(request, 'dashboard/upload_student_document.html', {'form': form, 'required_doc': required_doc})

@login_required
def coordinator_documents_overview(request):
    """Overview of all students' document completion (coordinator only)."""
    if not request.user.is_coordinator:
        messages.error(request, 'Only coordinators can view this overview.')
        return redirect('dashboard:home')
    from accounts.models import StudentProfile, RequiredDocument, StudentDocument
    students = StudentProfile.objects.select_related('user', 'course').all()
    required_documents = list(RequiredDocument.objects.all())
    students_with_docs = []
    for student in students:
        uploaded_docs = {doc.document_type_id: doc for doc in student.documents.all()}
        doc_status = []
        for req_doc in required_documents:
            doc = uploaded_docs.get(req_doc.id)
            doc_status.append({
                'required': req_doc,
                'uploaded': doc is not None,
                'file': doc.file.url if doc else None,
            })
        total = len(required_documents)
        completed = sum(1 for d in doc_status if d['uploaded'])
        progress = int((completed / total) * 100) if total > 0 else 0
        students_with_docs.append({
            'student': student,
            'doc_status': doc_status,
            'progress': progress,
        })
    context = {
        'students_with_docs': students_with_docs,
        'required_documents': required_documents,
    }
    return render(request, 'dashboard/coordinator_documents_overview.html', context)

@login_required
def coordinator_delete_student_document(request, student_id, doc_id):
    if not request.user.is_coordinator:
        messages.error(request, 'Only coordinators can delete student documents.')
        return redirect('dashboard:home')
    student = get_object_or_404(StudentProfile, id=student_id)
    required_doc = get_object_or_404(RequiredDocument, id=doc_id)
    doc = StudentDocument.objects.filter(student=student, document_type=required_doc).first()
    if doc:
        doc.file.delete(save=False)
        doc.delete()
        messages.success(request, f"Document '{required_doc.name}' deleted for {student.user.get_full_name()}.")
    else:
        messages.warning(request, 'Document not found.')
    return redirect('dashboard:coordinator_documents_overview')